"""
данный скрипт хорошо работает с современными отчетами тасс
с 2015 года по 2022 включительно
проверен 8 апреля 2022

"""

import os
import fnmatch
import os.path
import shutil
import openpyxl
from openpyxl import load_workbook
from selenium import webdriver
import time
import requests
from selenium.webdriver.common.by import By

report_dir = '/Volumes/big4photo/Downloads'
destination = '/Volumes/big4photo/Documents/TASS/reports'  # расположение обработанных файлов отчетов
main_report = '/Volumes/big4photo/Documents/TASS/Tass_total_report_from_2015.xlsx'  # файл куда сохранятеся вся

options = webdriver.ChromeOptions()
options.add_argument(
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4200.0 Iron Safari/537.36")


def gen_x(sheet):  # функция определяет номер строки с которой начинается ввод данных в табицу
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'компания' or cell.value == 'Компания':
                return cell.row + 1


def gen_y(sheet):  # функция определяет номер строки с которой начинается ввод данных в таблицу
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'ID фото ' or cell.value == 'ID фото':
                return cell.column


def get_prevue(file_to_work, report_date):
    browser = webdriver.Chrome(options=options)
    os.makedirs(f"/Volumes/big4photo/Documents/TASS/images/{report_date[1]}/{' '.join(report_date)}", exist_ok=True)
    wb = openpyxl.load_workbook(file_to_work)
    sheet = wb.active
    x = gen_x(sheet)
    y = gen_y(sheet)

    try:
        browser.get('https://www.tassphoto.com/ru')
        time.sleep(1)
        photo_id = (sheet.cell(row=x, column=y)).value
        while photo_id != None:
            search_input = browser.find_element(By.ID, "userrequest")
            search_input.clear()
            search_input.send_keys(photo_id)
            browser.find_element(By.ID, "search-submit").click()
            picture = browser.find_element(By.CSS_SELECTOR, "#mosaic .zoom img").get_attribute("src")
            print(picture)
            get_image = requests.get(picture)
            with open(
                    f"/Volumes/big4photo/Documents/TASS/images/{report_date[1]}/{' '.join(report_date)}/{photo_id}.jpg",
                    'wb') as img_file:
                img_file.write(get_image.content)
            x += 1
            photo_id = (sheet.cell(row=x, column=y)).value
    except Exception as ex:
        print(ex)
        browser.close()
        browser.quit()
    browser.close()
    browser.quit()


def write_to_main_file(photos, main_report, report_date):  # записываю информацию в главный файл
    wb = load_workbook(filename=main_report, read_only=False)
    ws_month_number = wb.create_sheet(" ".join(report_date), 0)
    ws_month_number.cell(row=1, column=1).value = "photo_id"
    ws_month_number.cell(row=1, column=2).value = "income"
    ws_month_number.cell(row=1, column=3).value = "sold times"

    kkeys = [i for i in photos.keys()]
    for i in range(len(photos)):
        ws_month_number.cell(row=2 + i, column=1).value = kkeys[i]
        ws_month_number.cell(row=2 + i, column=2).value = sum(photos[kkeys[i]])  # суммирую доход по снимкам
        ws_month_number.cell(row=2 + i, column=3).value = len(photos[kkeys[i]])
    wb.save(main_report)
    wb.close()
    print("информация записана")


def add_information_to_main_file(file_to_work,
                                 report_date):  # добавляю всю информацию из перенесенного файла отчета в основной файл
    wb = openpyxl.load_workbook(file_to_work)
    sheet = wb.active
    photos = {}  # словарь где ключ  id снимка , а значение список с цифрами покупок
    x = gen_x(sheet)  # номер строки с которого начинается отчет по фото
    y = gen_y(sheet)
    photo_id = (sheet.cell(row=x, column=y)).value
    money = float((sheet.cell(row=x, column=6)).value)
    while photo_id != None:
        photos.setdefault(photo_id, [])
        photos[photo_id].append(round(money, 2))
        x += 1
        photo_id = (sheet.cell(row=x, column=y)).value
        money = (sheet.cell(row=x, column=6)).value
    write_to_main_file(photos, main_report, report_date)


def get_report_date(file_name):  # получаю дату отчета в виде строки для новых отчетов точно с 2015
    wb = openpyxl.load_workbook(f"{report_dir}/{file_name}")
    sheet = wb.active
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Период:      ':
                return sheet.cell(row=cell.row, column=cell.column + 1).value


def move_and_rename(file_name, report_dir, destination):  # переименовываю и перемещаю файл отчета
    report_date = get_report_date(file_name).split()  # список с датой отчета [месяц, год, мусор]
    os.makedirs(f"{destination}/{report_date[1]}_отчеты", exist_ok=True)
    working_file = f"{destination}/{report_date[1]}_отчеты/Павленко_{report_date[0]}_{report_date[1]}.xlsx"
    if os.path.exists(
            working_file):  # надо вставить проверку на случай если файл отчета уже есть. то прекратить работу. удалив исходник
        os.remove(f"{report_dir}/{file_name}")
        print('данный отчет уже обработан ранее')
        return
    file_to_work = shutil.move(f"{report_dir}/{file_name}", working_file)
    print(f"обработан файл - {file_to_work}")  # главная переменная с которой дальше буду работать
    add_information_to_main_file(file_to_work, report_date)
    get_prevue(file_to_work, report_date)


def find_report(report_dir, destination):  # поск заданных файлов в папке загрузок
    list_of_files = os.listdir(report_dir)
    pattern = 'Павленко*.xlsx'
    count = 0
    for file_name in list_of_files:
        if fnmatch.fnmatch(file_name, pattern):
            count += 1
            move_and_rename(file_name, report_dir, destination)
    if count == 0:
        print('нет нужного файла')


find_report(report_dir, destination)
