from openpyxl import load_workbook


def write_to_main_file(photos, main_report, report_date):  # записываю информацию в главный файл
    wb = load_workbook(filename=main_report, read_only=False)
    ws_month_number = wb.create_sheet(" ".join(report_date), 0)
    ws_month_number.cell(row=1, column=1).value = "photo_id"
    ws_month_number.cell(row=1, column=2).value = "income"
    ws_month_number.cell(row=1, column=3).value = "sold times"

    kkeys = [i for i in photos.keys()]
    for i in range(len(photos)):
        ws_month_number.cell(row=2 + i, column=1).value = kkeys[i]
        ws_month_number.cell(row=2 + i, column=2).value = sum(photos[kkeys[i]])  # суммирую доход по снимкам
        ws_month_number.cell(row=2 + i, column=3).value = len(photos[kkeys[i]])
    wb.save(main_report)
    wb.close()
    print("информация записана")
