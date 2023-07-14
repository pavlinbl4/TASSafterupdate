from openpyxl import Workbook
import os
from datetime import datetime
from openpyxl import load_workbook

from must_have.make_documents_subfolder import make_documents_subfolder


def create_columns_names(ws):
    ws[f'A1'] = 'images_online'
    ws[f'B1'] = 'image_id'
    ws[f'C1'] = 'image_date'
    ws[f'D1'] = 'image_caption'
    ws[f'E1'] = 'image_link'


def create_xlsx(report_folder, file_name):
    today = datetime.now().strftime("%Y-%m-%d")
    if os.path.exists(f'{report_folder}/{file_name}.xlsx'):
        wb = load_workbook(
            f'{report_folder}/{file_name}.xlsx')  # файл есть и открываю его - too long to open big old file
        ws = wb.create_sheet(today)  # добавляю новую таблицу
        create_columns_names(ws)
    else:
        wb = Workbook()  # если файда еще нет
        ws = wb.active  # если файда еще нет
        ws.title = today  # если файда еще нет
        create_columns_names(ws)

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10  # задаю ширину колонки
    ws.column_dimensions['D'].width = 110
    ws.column_dimensions['E'].width = 50
    wb.save(file_name)
    # wb.close()
    return ws, wb


if __name__ == '__main__':
    report_folder = make_documents_subfolder('TASS/Tass_data')
    print(create_xlsx(report_folder, 'TASS_photo'))