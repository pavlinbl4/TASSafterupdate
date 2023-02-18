"""
сравниваю отчет по добавленным снимкам в TASS_photos.xlsx
и скачиваю превью снимков добавленых в самый последний раз
2022_03_07
"""

import pandas as pd
import datacompy
import os
import requests
import re
from openpyxl import load_workbook
from must_have.make_documents_subfolder import make_documents_subfolder


def add_data():  # из таблицы получаю дату из последней и предыдущей строки
    file = f"{report_folder}/TASS_photos.xlsx"
    book = load_workbook(file)
    ws = book.active
    last_row = ws.max_row
    last_date = f'{ws.cell(row=last_row, column=1).value[:10]}'
    print(f"{last_date = }")
    previous_date = f'{ws.cell(row=(last_row - 1), column=1).value[:10]}'
    print(f"{previous_date = }")
    difference = new_pictures_links(last_date,
                                    previous_date)  # вызываю функцию и получаю датафрэйм с новыми фото
    image_downloader(difference, last_date)  # скачиваю новые снимки


def image_downloader(difference, last_date):
    folder = f'[{report_folder}]/added_images'
    os.makedirs(f"{folder}/{last_date}", exist_ok=True)
    for i in range(len(difference)):
        image_url = difference.image_link.iloc[i]
        r = requests.get(image_url)
        image_name = re.findall(r'\d+(?=\.thw)', image_url)[0]
        print(image_name, image_url)
        with open(f"{folder}/{last_date}/{image_name}.jpg", 'wb') as download_file:
            for chunk in r.iter_content(9000):
                download_file.write(chunk)


def new_pictures_links(last_date, previous_date):
    pd.options.display.max_colwidth = 100
    photo_base = f'{report_folder}/all_TASS_images.xlsx'
    last_df = pd.read_excel(photo_base, sheet_name=last_date)
    previos_df = pd.read_excel(photo_base, sheet_name=previous_date)
    difference = datacompy.Compare(previos_df, last_df,
                                   join_columns=['image_id', 'image_date', 'image_caption']).df2_unq_rows

    return difference


if __name__ == "__main__":
    report_folder = make_documents_subfolder('TASS/Tass_data')
    add_data()
