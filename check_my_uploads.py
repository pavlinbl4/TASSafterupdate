"""проверяю, добавили ли мои фото на сайт, скрипт отрабатывает по crontab и
 заносит информацию в таблицу, если количество снимков изменилось, то запускаю скрипт
 all_images_new и потом добавляет свежие снимки в заданную папку """

from openpyxl import load_workbook
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook
import pandas as pd
import requests
import re
import datacompy

options = webdriver.ChromeOptions()
options.headless = True  # фоновый режим
options.add_argument(
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/"
    "537.36 (KHTML, like Gecko) Chrome/81.0.4200.0 Iron Safari/537.36")
options.add_argument("--disable-blink-features=AutomationControlled")

browser = webdriver.Chrome(options=options)

url = 'https://www.tassphoto.com/ru/asset/fullTextSearch/search/' \
      '%D0%A1%D0%B5%D0%BC%D0%B5%D0%BD%20%D0%9B%D0%B8%D1%85%D0%BE%D0%B4%D0%B5%D0%B5%D0%B2/page/'
report_folder = "/Volumes/big4photo/Documents/TASS/Tass_data"


def notification(message):
    title = "TASS info"
    command = f'''
    osascript -e 'display notification "{message}" with title "{title}"'
    '''
    os.system(command)


def create_columns_names(ws):
    ws[f'A1'] = 'images_online'
    ws[f'B1'] = 'image_id'
    ws[f'C1'] = 'image_date'
    ws[f'D1'] = 'image_caption'
    ws[f'E1'] = 'image_link'


def image_downloader(difference, last_date):
    folder = '/Volumes/big4photo/Documents/TASS/Tass_data/added_images'
    os.makedirs(f"{folder}/{last_date}", exist_ok=True)
    for i in range(len(difference)):
        image_url = difference.image_link.iloc[i]
        r = requests.get(image_url)
        image_name = re.findall(r'\d+(?=\.thw)', image_url)[0]
        print(image_name, image_url)
        with open(f"{folder}/{last_date}/{image_name}.jpg", 'wb') as download_file:
            for chunk in r.iter_content(9000):
                download_file.write(chunk)


def new_pictures_links(last_date, previous_date):
    pd.options.display.max_colwidth = 100
    photo_base = '/Volumes/big4photo/Documents/TASS/Tass_data/all_TASS_images.xlsx'
    last_df = pd.read_excel(photo_base, sheet_name=last_date)
    previos_df = pd.read_excel(photo_base, sheet_name=previous_date)
    difference = datacompy.Compare(previos_df, last_df,
                                   join_columns=['image_id', 'image_date', 'image_caption']).df2_unq_rows

    return difference


def create_xlsx():
    today = datetime.now().strftime("%Y-%m-%d")
    if os.path.exists(f'{report_folder}/all_TASS_images.xlsx'):
        wb = load_workbook(f'{report_folder}/all_TASS_images.xlsx')  # файл есть и открываю его
        ws = wb.create_sheet(today)  # добавляю новую таблицу
        create_columns_names(ws)
    else:
        wb = Workbook()  # если файда еще нет
        ws = wb.active  # если файда еще нет
        ws.title = today  # если файда еще нет
        create_columns_names(ws)

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10  # задаю ширину колонки
    ws.column_dimensions['D'].width = 110
    ws.column_dimensions['E'].width = 50
    return ws, wb


def check_all_images(page_number, images_online):  # 1. start to check images
    ws, wb = create_xlsx()
    count = 1
    for n in range(1, page_number + 1):  # количество страниц для анализа  - page_number + 1
        link = f'{url}{n}'
        soup = get_soup(get_html(link))
        thumbs_data = soup.find('ul', id="mosaic").find_all('div', class_="thumb-content thumb-width thumb-height")
        images_on_page = len(soup.find('ul', id="mosaic").find_all('a', class_="zoom"))
        for i in range(images_on_page):
            count += 1
            image_date = thumbs_data[i].find(class_="date").text
            image_id = thumbs_data[i].find(class_="title").text
            image_title = thumbs_data[i].find('p').text
            image_caption = soup.find('ul', id="mosaic").find_all(class_="thumb-text")[i].text.strip().split('\n')[
                -1].lstrip().replace(' Семен Лиходеев/ТАСС', '').replace(' Фото ИТАР-ТАСС/ Семен Лиходеев', '')
            image_link = soup.find('ul', id="mosaic").find_all('a', class_="zoom")[i].find('img').get('src')
            # print(count - 1, image_id, image_date)
            # print(image_title)
            # print(image_caption)
            # print(image_link)
            ws[f'A{count}'] = images_online + 1 - count
            ws[f'B{count}'] = image_id
            ws[f'C{count}'] = image_date
            ws[f'D{count}'] = image_caption
            ws[f'E{count}'] = image_link
    wb.save(f'{report_folder}/all_TASS_images.xlsx')
    wb.close()


def get_soup(html):
    soup = BeautifulSoup(html, 'lxml')
    return soup


def get_html(link):
    browser.get('https://www.tassphoto.com/ru')
    WebDriverWait(browser, 10).until(
        ec.presence_of_element_located((By.ID, "userrequest"))
    )
    search_input = browser.find_element(By.ID, "userrequest")
    search_input.clear()
    search_input.send_keys('Семен Лиходеев')
    search_input.send_keys(Keys.ENTER)
    browser.get(link)
    html = browser.page_source
    return html


def get_page_numbers():  # get number of images on site
    soup = get_soup(get_html(f'{url}1'))
    images_online = int(str(soup.select(".result-counter#nb-result"))[42:47])
    page_number = images_online // 20 + 1
    return page_number, images_online


def add_data():  # function check number of images
    page_number, images_online = get_page_numbers()
    # print(f'{images_online = }')
    file = "/Volumes/big4photo/Documents/TASS/Tass_data/TASS_photos.xlsx"
    book = load_workbook(file)
    ws = book.active
    last_row = ws.max_row
    old_value = ws.cell(row=last_row, column=2).value  # последние данные в таблице
    # print(f'date of last row {ws.cell(row=last_row, column=1).value[:10]} - images {old_value}')
    if ws.cell(row=last_row, column=2).value != images_online:
        ws.cell(row=last_row + 1, column=2).value = images_online
        ws.cell(row=last_row + 1, column=1).value = datetime.now().strftime('%Y-%m-%d %H:%M')
        new_images = int(images_online) - int(old_value)  # количество добавленных фото
        ws.cell(row=last_row + 1, column=3).value = new_images
        print(f'добавлено {new_images} снимков')
        notification(f'добавлено\n{new_images}\nснимков')
        book.save(file)
        book.close()

        check_all_images(page_number, images_online)  # create new sheet with images on site

        # dates of last and previous images update
        book = load_workbook(file)
        ws = book.active
        last_row = ws.max_row
        last_date = f'{ws.cell(row=last_row, column=1).value[:10]}'
        print(f"{last_date = }")
        previous_date = f'{ws.cell(row=(last_row - 1), column=1).value[:10]}'
        # print(f"{previous_date = }")

        difference = new_pictures_links(last_date,
                                        previous_date)  # вызываю функцию и получаю датафрэйм с новыми фото

        image_downloader(difference, last_date)  # скачиваю новые снимки

    else:
        print(f"no new images, now  {images_online} - images online")
        notification(f"no new images\n{images_online}\nimages online")

    book.save(file)
    book.close()
    print("work completed")


add_data()
