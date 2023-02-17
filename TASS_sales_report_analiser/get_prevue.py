import time
import os
from selenium import webdriver
import openpyxl
import requests
from selenium.webdriver.common.by import By
from TASS_sales_report_analiser.home_directory import subfolder_in_user_folder
from TASS_sales_report_analiser.sale_report_analiser import gen_x, gen_y
from crome_options import setting_chrome_options


def get_prevue(file_to_work, report_date):
    browser = webdriver.Chrome(options=setting_chrome_options())
    os.makedirs(f"{subfolder_in_user_folder('Documents')}/TASS/images/{report_date[1]}/{' '.join(report_date)}",
                exist_ok=True)
    wb = openpyxl.load_workbook(file_to_work)
    sheet = wb.active
    x = gen_x(sheet)
    y = gen_y(sheet)

    try:
        browser.get('https://www.tassphoto.com/ru')
        time.sleep(1)
        photo_id = (sheet.cell(row=x, column=y)).value
        while photo_id is not None:
            search_input = browser.find_element(By.ID, "userrequest")
            search_input.clear()
            search_input.send_keys(photo_id)
            browser.find_element(By.ID, "search-submit").click()
            picture = browser.find_element(By.CSS_SELECTOR, f"img.thumb{photo_id}").get_attribute("src")
            print(picture)
            get_image = requests.get(picture)
            with open(
                    f"{subfolder_in_user_folder('Documents')}/TASS/images/{report_date[1]}"
                    f"/{' '.join(report_date)}/{photo_id}.jpg",
                    'wb') as img_file:
                img_file.write(get_image.content)
            x += 1
            photo_id = (sheet.cell(row=x, column=y)).value
    except Exception as ex:
        print(ex)
        browser.close()
        browser.quit()
    browser.close()
    browser.quit()


if __name__ == '__main__':

    get_prevue('/Users/evgeniy/Documents/TASS/reports/2023_отчеты/test_report.xlsx',
                ['month_name','YYEAR', 'года'])
