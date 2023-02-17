from openpyxl import load_workbook
from openpyxl import Workbook
import os
from must_have.make_documents_subfolder import make_documents_subfolder


def create_report_file(report_folder, file_name, report_date):
    report_file_name = f"{file_name}.xlsx"
    path_to_file = f'{report_folder}/{report_file_name}'

    if os.path.exists(path_to_file):
        wb = load_workbook(path_to_file)  # файл есть и открываю его
        ws = wb.create_sheet(" ".join(report_date))  # добавляю новую таблицу
    else:
        wb = Workbook()  # если файда еще нет
        ws = wb.active  # если файа еще нет
        ws.title = " ".join(report_date)  # если файда еще нет

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30  # задаю ширину колонки

    ws['A1'] = 'photo_id'  # create columns names
    ws['B1'] = 'income'
    ws['C1'] = 'sold times'

    wb.save(path_to_file)

    return path_to_file


def create_report(report_date):
    return create_report_file(make_documents_subfolder("TASS"), "all_years_report", report_date)


if __name__ == '__main__':
    create_report(['январь', '2023', 'года'])
