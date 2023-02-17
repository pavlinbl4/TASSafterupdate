import os
import fnmatch
import os.path
import shutil
import openpyxl
from openpyxl import load_workbook
from get_preview import get_preview
from must_have.home_directory import subfolder_in_user_folder
from create_XLXS_report_file import create_report
from read_XLSX_report import gen_x, gen_y



def write_to_main_file(photos, main_report, report_date):  # записываю информацию в главный файл
    wb = load_workbook(filename=main_report, read_only=False)
    ws_month_number = wb.create_sheet(" ".join(report_date), 0)
    ws_month_number.cell(row=1, column=1).value = "photo_id"
    ws_month_number.cell(row=1, column=2).value = "income"
    ws_month_number.cell(row=1, column=3).value = "sold times"

    kkeys = [i for i in photos.keys()]
    for i in range(len(photos)):
        ws_month_number.cell(row=2 + i, column=1).value = kkeys[i]
        ws_month_number.cell(row=2 + i, column=2).value = sum(photos[kkeys[i]])  # суммирую доход по снимкам
        ws_month_number.cell(row=2 + i, column=3).value = len(photos[kkeys[i]])
    wb.save(main_report)
    wb.close()
    print("информация записана")


def add_information_to_main_file(file_to_work,
                                 report_date):  # добавляю всю информацию из перенесенного файла отчета в основной файл

    wb = openpyxl.load_workbook(file_to_work)
    sheet = wb.active
    photos = {}  # словарь где ключ  id снимка , а значение список с цифрами покупок
    x = gen_x(sheet)  # номер строки с которого начинается отчет по фото
    y = gen_y(sheet)
    photo_id = (sheet.cell(row=x, column=y)).value
    money = float((sheet.cell(row=x, column=6)).value)
    while photo_id is not None:
        photos.setdefault(photo_id, [])
        photos[photo_id].append(round(money, 2))
        x += 1
        photo_id = (sheet.cell(row=x, column=y)).value
        money = (sheet.cell(row=x, column=6)).value
    main_report = create_report(report_date)

    write_to_main_file(photos, main_report, report_date)


def get_report_date(file_name, report_dir):  # получаю дату отчета в виде строки для новых отчетов точно с 2015
    wb = openpyxl.load_workbook(f"{report_dir}/{file_name}")
    sheet = wb.active
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Период:      ':
                return sheet.cell(row=cell.row, column=cell.column + 1).value


def move_and_rename(file_name, report_dir, destination):  # переименовываю и перемещаю файл отчета
    report_date = get_report_date(file_name, report_dir).split()  # список с датой отчета [месяц, год, мусор]
    os.makedirs(f"{destination}/{report_date[1]}_отчеты", exist_ok=True)
    working_file = f"{destination}/{report_date[1]}_отчеты/Павленко_{report_date[0]}_{report_date[1]}.xlsx"
    if os.path.exists(
            working_file):  # надо вставить проверку на случай
        # если файл отчета уже есть. то прекратить работу. удалив исходник
        os.remove(f"{report_dir}/{file_name}")
        print('данный отчет уже обработан ранее')
        exit()
    return shutil.move(f"{report_dir}/{file_name}", working_file), report_date  # file_to_work, report_date


def find_report(report_dir, destination):  # поск заданных файлов в папке загрузок
    list_of_files = os.listdir(report_dir)
    pattern = 'Павленко*.xlsx'
    count = 0
    for file_name in list_of_files:
        if fnmatch.fnmatch(file_name, pattern):
            count += 1
            file_to_work = move_and_rename(file_name, report_dir, destination)
            return file_to_work
    if count == 0:
        print('В папке загрузки нет файла отчета')
        exit()


def main():
    report_dir = subfolder_in_user_folder('Downloads')
    tass_folder = f'{subfolder_in_user_folder("Documents")}/TASS'
    destination = f'{tass_folder}/reports'  # расположение обработанных файлов отчетов

    file_to_work, report_date = find_report(report_dir, destination)

    print(f"обработан файл - {file_to_work}")  # главная переменная с которой дальше буду работать

    add_information_to_main_file(file_to_work, report_date)
    get_preview(file_to_work, report_date)


if __name__ == '__main__':
    main()
