import time
import os
from selenium import webdriver
import openpyxl
import requests
from selenium.webdriver.common.by import By
from home_directory import subfolder_in_user_folder
from read_XLSX_report import gen_x, gen_y
from crome_options import setting_chrome_options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec


def get_preview(file_to_work, report_date):
    browser = webdriver.Chrome(options=setting_chrome_options())
    os.makedirs(f"{subfolder_in_user_folder('Documents')}/TASS/images/{report_date[1]}/{' '.join(report_date)}",
                exist_ok=True)
    wb = openpyxl.load_workbook(file_to_work)
    sheet = wb.active
    x = gen_x(sheet)
    y = gen_y(sheet)

    try:
        browser.get('https://www.tassphoto.com/ru')
        # time.sleep(3)
        WebDriverWait(browser, 10).until(
            ec.presence_of_element_located((By.ID, "userrequest"))
        )
        photo_id = (sheet.cell(row=x, column=y)).value
        while photo_id is not None:
            # search_input = browser.find_element(By.ID, "userrequest")
            # search_input.clear()
            # search_input.send_keys(photo_id)
            browser.get(f'https://www.tassphoto.com/ru/asset/fullTextSearch/search/{photo_id}/page/1')
            # browser.find_element(By.ID, "search-submit").click()
            WebDriverWait(browser, 10).until(
                ec.presence_of_element_located((By.ID, "userrequest"))
            )

            picture = browser.find_element(By.CSS_SELECTOR, f"img.thumb{photo_id}").get_attribute("src")
            print(picture)
            get_image = requests.get(picture)
            with open(
                    f"{subfolder_in_user_folder('Documents')}/TASS/images/{report_date[1]}"
                    f"/{' '.join(report_date)}/{photo_id}.jpg",
                    'wb') as img_file:
                img_file.write(get_image.content)
            x += 1
            photo_id = (sheet.cell(row=x, column=y)).value
    except Exception as ex:
        print(ex)
        browser.close()
        browser.quit()
    browser.close()
    browser.quit()


if __name__ == '__main__':
    get_preview('/Users/evgeniy/Documents/TASS/reports/2023_отчеты/Павленко_январь_2023.xlsx',
                ['month_name', 'YYEAR', 'года'])
